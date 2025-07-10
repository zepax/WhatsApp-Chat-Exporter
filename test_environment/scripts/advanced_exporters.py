"""
Advanced Export Functions for WhatsApp Chat Analyzer
Supports Excel, PDF, and enhanced reporting formats
"""

import json
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
import logging

def generate_advanced_excel_report(analyzer_instance, output_dir: str = "analysis_results"):
    """Generate comprehensive Excel report with multiple sheets and charts."""
    if not EXCEL_ADVANCED:
        logging.warning("Advanced Excel features not available. Install openpyxl for full functionality.")
        return None
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create workbook
        wb = Workbook()
        
        # Define styles
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        title_font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
        normal_font = Font(name='Calibri', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Sheet 1: Executive Summary
        ws_summary = wb.active
        ws_summary.title = "📊 Resumen Ejecutivo"
        
        # Summary title
        ws_summary['A1'] = "REPORTE DE ANÁLISIS DE CONVERSACIONES WHATSAPP"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:F1')
        
        # Summary statistics
        summary = analyzer_instance.generate_summary_report()
        
        row = 3
        summary_data = [
            ("📅 Fecha del análisis:", datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("📁 Total de archivos analizados:", f"{summary.get('total_files_analyzed', 0):,}"),
            ("💬 Total de mensajes:", f"{summary.get('total_messages', 0):,}"),
            ("📝 Total de palabras:", f"{summary.get('total_words', 0):,}"),
            ("🎯 Archivos con coincidencias:", f"{summary.get('files_with_keyword_matches', 0):,}"),
            ("⚡ Promedio mensajes por archivo:", f"{summary.get('average_messages_per_file', 0):.1f}"),
            ("📊 Promedio palabras por archivo:", f"{summary.get('average_words_per_file', 0):.1f}")
        ]
        
        for label, value in summary_data:
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Top keywords section
        ws_summary[f'A{row + 1}'] = "🔍 PALABRAS CLAVE MÁS FRECUENTES"
        ws_summary[f'A{row + 1}'].font = title_font
        row += 3
        
        ws_summary[f'A{row}'] = "Palabra Clave"
        ws_summary[f'B{row}'] = "Frecuencia"
        ws_summary[f'C{row}'] = "Porcentaje"
        
        for col in ['A', 'B', 'C']:
            ws_summary[f'{col}{row}'].font = header_font
            ws_summary[f'{col}{row}'].fill = header_fill
            ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center')
        
        top_keywords = summary.get('top_keywords', {})
        total_keyword_count = sum(top_keywords.values())
        
        row += 1
        for keyword, count in list(top_keywords.items())[:10]:
            percentage = (count / total_keyword_count * 100) if total_keyword_count > 0 else 0
            ws_summary[f'A{row}'] = keyword
            ws_summary[f'B{row}'] = count
            ws_summary[f'C{row}'] = f"{percentage:.1f}%"
            row += 1
        
        # Sheet 2: Detailed Results
        ws_details = wb.create_sheet("📋 Resultados Detallados")
        
        # Headers
        headers = ["Archivo", "Mensajes", "Palabras", "Coincidencias", "Sentimiento", 
                  "Idioma", "Participantes", "Temas", "Tamaño (KB)", "Tiempo Procesamiento"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Data rows
        for row_idx, result in enumerate(analyzer_instance.results, 2):
            data = [
                result.file_name,
                result.total_messages,
                result.total_words,
                sum(result.keyword_matches.values()),
                getattr(result, 'sentiment_score', 'N/A'),
                getattr(result, 'language_detected', 'N/A'),
                ', '.join(getattr(result, 'participants', [])),
                ', '.join(getattr(result, 'topics_detected', [])),
                getattr(result, 'file_size_kb', 0),
                f"{getattr(result, 'processing_time', 0):.2f}s"
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_details.cell(row=row_idx, column=col, value=value)
                cell.font = normal_font
                cell.border = border
                if col in [2, 3, 4, 9, 10]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for column in ws_details.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_details.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 3: Keywords Analysis
        ws_keywords = wb.create_sheet("🔍 Análisis de Palabras")
        
        # Create keyword matrix
        all_keywords = set()
        for result in analyzer_instance.results:
            all_keywords.update(result.keyword_matches.keys())
        
        all_keywords = sorted(list(all_keywords))
        
        # Headers
        ws_keywords['A1'] = "Archivo"
        ws_keywords['A1'].font = header_font
        ws_keywords['A1'].fill = header_fill
        
        for col, keyword in enumerate(all_keywords, 2):
            cell = ws_keywords.cell(row=1, column=col, value=keyword)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', text_rotation=45)
        
        # Data matrix
        for row_idx, result in enumerate(analyzer_instance.results, 2):
            ws_keywords.cell(row=row_idx, column=1, value=result.file_name)
            
            for col, keyword in enumerate(all_keywords, 2):
                count = result.keyword_matches.get(keyword, 0)
                cell = ws_keywords.cell(row=row_idx, column=col, value=count)
                
                # Color coding based on frequency
                if count > 0:
                    if count >= 10:
                        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                    elif count >= 5:
                        cell.fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
                    elif count >= 1:
                        cell.fill = PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid')
        
        # Sheet 4: Charts and Visualization
        ws_charts = wb.create_sheet("📈 Gráficos")
        
        # Create bar chart for top keywords
        if top_keywords:
            chart_data = []
            chart_keywords = list(top_keywords.keys())[:10]
            chart_counts = [top_keywords[k] for k in chart_keywords]
            
            # Add data to sheet
            ws_charts['A1'] = "Palabra Clave"
            ws_charts['B1'] = "Frecuencia"
            
            for i, (keyword, count) in enumerate(zip(chart_keywords, chart_counts), 2):
                ws_charts[f'A{i}'] = keyword
                ws_charts[f'B{i}'] = count
            
            # Create chart
            chart = BarChart()
            chart.title = "Top 10 Palabras Clave"
            chart.x_axis.title = "Palabras Clave"
            chart.y_axis.title = "Frecuencia"
            
            data_ref = Reference(ws_charts, min_col=2, min_row=1, max_row=len(chart_keywords) + 1)
            cats_ref = Reference(ws_charts, min_col=1, min_row=2, max_row=len(chart_keywords) + 1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            ws_charts.add_chart(chart, "D2")
        
        # Save workbook
        output_path = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_path / f"whatsapp_analysis_report_{timestamp}.xlsx"
        
        wb.save(excel_file)
        
        logging.info(f"Advanced Excel report generated: {excel_file}")
        return excel_file
        
    except Exception as e:
        logging.error(f"Failed to generate Excel report: {e}")
        return None

def generate_pdf_report(analyzer_instance, output_dir: str = "analysis_results"):
    """Generate professional PDF report with charts and analysis."""
    if not PDF_AVAILABLE:
        logging.warning("PDF generation not available. Install reportlab for PDF functionality.")
        return None
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.charts.piecharts import Pie
        
        # Setup
        output_path = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_file = output_path / f"whatsapp_analysis_report_{timestamp}.pdf"
        
        # Create document
        doc = SimpleDocTemplate(str(pdf_file), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#1F4E79'),
            alignment=1  # Center
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor('#366092')
        )
        
        # Title
        story.append(Paragraph("📊 REPORTE DE ANÁLISIS DE CONVERSACIONES WHATSAPP", title_style))
        story.append(Spacer(1, 12))
        
        # Summary section
        summary = analyzer_instance.generate_summary_report()
        
        story.append(Paragraph("📋 Resumen Ejecutivo", subtitle_style))
        
        summary_text = f"""
        <b>Fecha del análisis:</b> {datetime.now().strftime("%d/%m/%Y %H:%M")}<br/>
        <b>Total de archivos analizados:</b> {summary.get('total_files_analyzed', 0):,}<br/>
        <b>Total de mensajes:</b> {summary.get('total_messages', 0):,}<br/>
        <b>Total de palabras:</b> {summary.get('total_words', 0):,}<br/>
        <b>Archivos con coincidencias:</b> {summary.get('files_with_keyword_matches', 0):,}<br/>
        <b>Promedio mensajes por archivo:</b> {summary.get('average_messages_per_file', 0):.1f}<br/>
        <b>Promedio palabras por archivo:</b> {summary.get('average_words_per_file', 0):.1f}
        """
        
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Top keywords table
        story.append(Paragraph("🔍 Palabras Clave Más Frecuentes", subtitle_style))
        
        top_keywords = summary.get('top_keywords', {})
        if top_keywords:
            keyword_data = [['Palabra Clave', 'Frecuencia', 'Porcentaje']]
            total_count = sum(top_keywords.values())
            
            for keyword, count in list(top_keywords.items())[:10]:
                percentage = (count / total_count * 100) if total_count > 0 else 0
                keyword_data.append([keyword, f"{count:,}", f"{percentage:.1f}%"])
            
            keyword_table = Table(keyword_data, colWidths=[2*inch, 1*inch, 1*inch])
            keyword_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(keyword_table)
        
        story.append(Spacer(1, 20))
        
        # Conversation breakdown
        story.append(Paragraph("💬 Desglose por Conversación", subtitle_style))
        
        conv_breakdown = summary.get('conversation_breakdown', [])
        if conv_breakdown:
            conv_data = [['Archivo', 'Mensajes', 'Palabras', 'Coincidencias']]
            
            for conv in conv_breakdown[:15]:  # Top 15 conversations
                conv_data.append([
                    conv['file_name'][:30] + '...' if len(conv['file_name']) > 30 else conv['file_name'],
                    f"{conv['total_messages']:,}",
                    f"{conv['total_words']:,}",
                    f"{conv['keyword_count']:,}"
                ])
            
            conv_table = Table(conv_data, colWidths=[3*inch, 1*inch, 1*inch, 1*inch])
            conv_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8)
            ]))
            
            story.append(conv_table)
        
        # Add page break
        story.append(PageBreak())
        
        # Technical details
        story.append(Paragraph("🔧 Detalles Técnicos", subtitle_style))
        
        tech_details = f"""
        <b>Configuración del análisis:</b><br/>
        • Palabras clave monitoreadas: {len(analyzer_instance.config.keywords)}<br/>
        • Análisis de sentimientos: {'Activado' if analyzer_instance.config.sentiment_analysis else 'Desactivado'}<br/>
        • Detección de idioma: {'Activada' if analyzer_instance.config.language_detection else 'Desactivada'}<br/>
        • Extracción de temas: {'Activada' if analyzer_instance.config.topic_extraction else 'Desactivada'}<br/>
        • Uso de caché: {'Activado' if analyzer_instance.config.use_cache else 'Desactivado'}<br/>
        • Procesamiento paralelo: {analyzer_instance.config.max_workers} workers<br/><br/>
        
        <b>Información del sistema:</b><br/>
        • Versión del analizador: 2.0 Professional<br/>
        • Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}<br/>
        • Tiempo total de procesamiento: {sum(getattr(r, 'processing_time', 0) for r in analyzer_instance.results):.2f} segundos
        """
        
        story.append(Paragraph(tech_details, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        logging.info(f"PDF report generated: {pdf_file}")
        return pdf_file
        
    except Exception as e:
        logging.error(f"Failed to generate PDF report: {e}")
        return None

def generate_json_export(analyzer_instance, output_dir: str = "analysis_results", include_raw_data: bool = False):
    """Generate comprehensive JSON export with all analysis data."""
    try:
        output_path = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_file = output_path / f"whatsapp_analysis_export_{timestamp}.json"
        
        # Prepare export data
        export_data = {
            "metadata": {
                "export_version": "2.0",
                "generated_at": datetime.now().isoformat(),
                "analyzer_config": {
                    "keywords": analyzer_instance.config.keywords,
                    "use_regex": analyzer_instance.config.use_regex,
                    "case_sensitive": analyzer_instance.config.case_sensitive,
                    "sentiment_analysis": analyzer_instance.config.sentiment_analysis,
                    "language_detection": analyzer_instance.config.language_detection,
                    "topic_extraction": analyzer_instance.config.topic_extraction
                },
                "processing_stats": {
                    "total_files": len(analyzer_instance.results),
                    "total_processing_time": sum(getattr(r, 'processing_time', 0) for r in analyzer_instance.results),
                    "cache_hits": getattr(analyzer_instance, 'cache_hits', 0),
                    "errors_encountered": sum(len(getattr(r, 'errors', [])) for r in analyzer_instance.results)
                }
            },
            "summary": analyzer_instance.generate_summary_report(),
            "detailed_results": []
        }
        
        # Add detailed results
        for result in analyzer_instance.results:
            result_data = {
                "file_info": {
                    "name": result.file_name,
                    "path": getattr(result, 'file_path', ''),
                    "size_kb": getattr(result, 'file_size_kb', 0),
                    "hash": getattr(result, 'file_hash', ''),
                },
                "content_analysis": {
                    "total_messages": result.total_messages,
                    "total_words": result.total_words,
                    "keyword_matches": result.keyword_matches,
                    "participants": getattr(result, 'participants', []),
                    "date_range": getattr(result, 'date_range', [None, None]),
                    "message_frequency": getattr(result, 'message_frequency', {}),
                    "media_count": getattr(result, 'media_count', {})
                },
                "advanced_analysis": {
                    "sentiment_score": getattr(result, 'sentiment_score', None),
                    "language_detected": getattr(result, 'language_detected', None),
                    "topics_detected": getattr(result, 'topics_detected', []),
                },
                "processing_info": {
                    "processing_time": getattr(result, 'processing_time', 0),
                    "timestamp": result.timestamp,
                    "errors": getattr(result, 'errors', [])
                }
            }
            
            export_data["detailed_results"].append(result_data)
        
        # Write JSON file
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
        
        logging.info(f"JSON export generated: {json_file}")
        return json_file
        
    except Exception as e:
        logging.error(f"Failed to generate JSON export: {e}")
        return None

def generate_csv_export(analyzer_instance, output_dir: str = "analysis_results"):
    """Generate CSV export for data analysis tools."""
    if not PANDAS_AVAILABLE:
        logging.warning("CSV export requires pandas. Install pandas for full functionality.")
        return None
    
    try:
        import pandas as pd
        
        output_path = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_file = output_path / f"whatsapp_analysis_data_{timestamp}.csv"
        
        # Prepare data for DataFrame
        data_rows = []
        
        for result in analyzer_instance.results:
            row = {
                'file_name': result.file_name,
                'file_size_kb': getattr(result, 'file_size_kb', 0),
                'total_messages': result.total_messages,
                'total_words': result.total_words,
                'keyword_matches_count': sum(result.keyword_matches.values()),
                'sentiment_score': getattr(result, 'sentiment_score', None),
                'language_detected': getattr(result, 'language_detected', None),
                'participants_count': len(getattr(result, 'participants', [])),
                'topics_count': len(getattr(result, 'topics_detected', [])),
                'processing_time': getattr(result, 'processing_time', 0),
                'has_errors': len(getattr(result, 'errors', [])) > 0
            }
            
            # Add individual keyword counts
            for keyword in analyzer_instance.config.keywords:
                row[f'keyword_{keyword}'] = result.keyword_matches.get(keyword, 0)
            
            data_rows.append(row)
        
        # Create DataFrame and save
        df = pd.DataFrame(data_rows)
        df.to_csv(csv_file, index=False, encoding='utf-8')
        
        logging.info(f"CSV export generated: {csv_file}")
        return csv_file
        
    except Exception as e:
        logging.error(f"Failed to generate CSV export: {e}")
        return None